import openpyxl

# load_workbook is to read our file and loading file to a variable
inventory_file = openpyxl.load_workbook("inventory.xlsx")

# To access a certain sheet from the file such as sheet1 or product sheet etc
product_list = inventory_file["Sheet1"]
# print(product_list.max_row)

# TASK-1:
# Calculate the number of products per supplier and list
prod_per_supplier = {}  # Task-1 Dictionary
total_value_per_supplier = {}  # Task-2 Dictionary
products_under_ten = {}  # Task-3 Dictionary


# Run a loop to iterate through the file and do the necessary
for prod_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(prod_row, 4).value
    # Extract price and product from file
    inventory = product_list.cell(prod_row, 2).value
    price = product_list.cell(prod_row, 3).value
    product_num = product_list.cell(prod_row, 1).value
    # To add a new column and change the file
    inv_price = product_list.cell(prod_row, 5)

    # To calculate the products per supplier
    if supplier_name in prod_per_supplier:
        # current_number = prod_per_supplier[supplier_name]
        prod_per_supplier[supplier_name] = prod_per_supplier.get(supplier_name)
        prod_per_supplier[supplier_name] = prod_per_supplier[supplier_name] + 1
    else:
        print("Added a new supplier")
        prod_per_supplier[supplier_name] = 1

    # TASK-2 To calculate the total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Task-3: To print the products who which are less than 10 in stock or inventory
    if inventory < 10:
        products_under_ten[product_num] = inventory

#     Task-4: add value for total inventory price
    inv_price.value = inventory * price

print(prod_per_supplier)
print(total_value_per_supplier)
print(products_under_ten)

inventory_file.save("Updated_Inventory_file")
